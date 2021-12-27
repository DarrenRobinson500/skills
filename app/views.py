from django.shortcuts import render, redirect
from django.db.models import Avg
from .models import *
from .forms import *
from .filters import *
import openpyxl as xl

def notes_list(request):
    items = Note.objects.filter(level=1)
    return render(request, 'notes_list.html', {'items': items})

def calendar(request):
    items = Note.objects.exclude(date__isnull=True).exclude(type="Meeting").exclude(type="Person").exclude(status="Complete").order_by('date')
    bdays = Note.objects.filter(type="Person").exclude(date__isnull=True)
    items = items | bdays
    items = sorted(items, key=lambda t: t.calendar_date())

    return render(request, 'calendar.html', {'items': items})

def get_form(request, type, item):
    if type is None: type = "Note"

    if type == "Note": form = BaseForm(request.POST or None, instance=item)
    if type == "Person": form = PersonForm(request.POST or None, instance=item)
    if type == "Objective": form = ObjectiveForm(request.POST or None, instance=item)
    if type == "Story": form = StoryForm(request.POST or None, instance=item)
    if type == "Issue": form = IssueForm(request.POST or None, instance=item)
    if type == "ToDo": form = ToDoForm(request.POST or None, instance=item)
    if type == "Group": form = GroupForm(request.POST or None, instance=item)
    if type == "Reminder": form = ReminderForm(request.POST or None, instance=item)
    if type == "Meeting": form = MeetingForm(request.POST or None, instance=item)

    return form

def new(request, type=None, parent_id=None, return_page=None):
    if type is None: type = "Note"
    form = get_form(request, type, None)
    heading = "New "
    if form.is_valid():
        new = form.save()
        if type is not None: new.type = type
        if parent_id is not None:
            parent = Note.objects.get(id=parent_id)
            new.parent = parent
            new.level = parent.level + 1
        else:
            new.level = 1
        new.save()
        if return_page is None: return redirect("notes_list")
        return redirect("/ind/" + str(return_page))
    return render(request, "new.html", {"type": type, "heading": heading, "form": form})

def delete(request,id):
    item = Note.objects.get(id=id)
    parent = item.parent
    item.delete()
    if parent is None: return(redirect("notes_list"))
    return redirect("/ind/" + str(parent.id))

def complete(request,id):
    item = Note.objects.get(id=id)
    item.status = "Complete"
    item.save()
    return redirect("calendar")



def get_types(type):
    types = None
    if type == "Person":
        # type, Table Heading
        types = [("ToDo", "ToDo"), ("Person", "Related People"), ("Objective", "Objectives"), ("Story", "Stories"), ("Issue", "Issues"),
                 ("ToDo", "ToDo"), ("Group", "Groups"), ("Reminder", "Reminders"), ("Meeting", "Meetings"), ]
    if type == "Group":
        types = [("ToDo", "ToDo"), ("Person", "Related People"), ("Objective", "Objectives"), ("Story", "Stories"), ("Issue", "Issues"),
                 ("Group", "Groups"), ("Reminder", "Reminders"), ("Meeting", "Meetings"), ]
    if type == "Objective":
        types = [("ToDo", "ToDo"), ("Objective", "Objectives"), ("Issue", "Issues"), ("Reminder", "Reminders"), ("Meeting", "Meetings"), ]
    return types

def ind(request, id):
    item = Note.objects.get(id=id)
    form = get_form(request, item.type, item)
    types = get_types(item.type)
    if form.is_valid():
        form.save()

    return render(request, "ind.html", {"item": item, "types": types, "form": form})

def parent(request, id):
    item = Note.objects.get(id=id)
    form = ParentForm(request.POST or None, instance=item)
    heading = "Change parent of "
    if form.is_valid():
        form.save()
        return redirect("/ind/" + str(id))


    return render(request, "new.html", {"item": item, "heading": heading, "form": form})



def people_list(request):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    list = People.objects.all().order_by('name')
    # filter = PeopleFilter()
    filter = None
    return render(request, 'people_list.html', {'list': list, 'filter':filter, 'colour':colour})

def people_team(request, id):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    person = People.objects.get(id=id)
    team1 = People.objects.filter(manager=person)
    team2 = People.objects.none()
    for member in team1:
        team2 = team2 | People.objects.filter(manager=member)
    team3 = People.objects.none()
    for member in team2:
        team3 = team3 | People.objects.filter(manager=member)
    team4 = People.objects.none()
    for member in team3:
        team4 = team4 | People.objects.filter(manager=member)
    team5 = People.objects.none()
    for member in team4:
        team5 = team5 | People.objects.filter(manager=member)
    team = team1 | team2 | team3 | team4 | team5

    skills = Skill.objects.all()
    levels = Skill_Level.objects.all()


    for skill in skills:
        skill.text0_temp = ""
        skill.text1_temp = ""
        skill.text2_temp = ""
        skill.text3_temp = ""
        skill.text4_temp = ""
        skill.text5_temp = ""

    for member in team:
        for skill in skills:
            score_obj = Score.objects.filter(person=member,skill=skill)
            if score_obj.count() == 0:
                score = 0
            else:
                score_obj = score_obj[0]
                score = score_obj.score
            if score == 0: skill.text0_temp += member.name + chr(13)
            if score == 1: skill.text1_temp += member.name + chr(13)
            if score == 2: skill.text2_temp += member.name + chr(13)
            if score == 3: skill.text3_temp += member.name + chr(13)
            if score == 4: skill.text4_temp += member.name + chr(13)
            if score == 5: skill.text5_temp += member.name + chr(13)
            skill.save()

    for skill in skills:
        print(skill.question)


    return render(request, "people_team.html", {'person':person, 'team':team, 'skills': skills, 'levels':levels, 'colour':colour})



def people_ind(request, id):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    list = Skill.objects.all()
    levels = Skill_Level.objects.all()
    sub_categories = Skill_Cat.objects.all()
    person = People.objects.get(id=id)
    scores = Score.objects.filter(person=person)
    for sub_category in sub_categories:
        sub_category.to_develop = ""
        skills = Skill.objects.filter(sub_category=sub_category)
        for skill in skills:
            if Score.objects.filter(person=person, skill=skill).count() > 0:
                result = Score.objects.filter(person=person, skill=skill)[0].score
                print(skill.id,skill.question,person.level(),skill.level(),result)
                if person.level() == skill.level() and result < 3:
                    sub_category.to_develop += skill.question + chr(13)
                if person.level() == skill.level() + 1 and result < 4:
                    sub_category.to_develop += skill.question + " (" + levels[3].level + ")"+ chr(13)
                if person.level() == skill.level() - 1 and result < 2:
                    sub_category.to_develop += skill.question + " (" + levels[1].level + ")" + chr(13)
            else:
                sub_category.to_develop += skill.question + " [No response]" + chr(13)
        sub_category.save()

        # for sub_category in sub_categories:
    #     sub_category.score1 = 0.0
    #     sub_category.score2 = 0.0
    #     sub_category.score3 = 0.0
    #     sub_category.score4 = 0.0
    #     role_levels = Role_Level.objects.all()
    #     for count,role_level in enumerate(role_levels,start=1):
    #         if Skill.objects.filter(sub_category=sub_category,role_level=role_level).count() > 0:
    #             skills = Skill.objects.filter(sub_category=sub_category,role_level=role_level)
    #             skill_count = 0
    #             total_score = 0
    #             for skill in skills:
    #                 if Score.objects.filter(person=person, skill=skill).count() > 0:
    #                     total_score += Score.objects.filter(person=person,skill=skill)[0].score
    #                     skill_count += 1
    #             if skill_count > 0:
    #                 if count == 1: sub_category.score1 = round(total_score / skill_count,1)
    #                 if count == 2: sub_category.score2 = round(total_score / skill_count,1)
    #                 if count == 3: sub_category.score3 = round(total_score / skill_count,1)
    #                 if count == 4: sub_category.score4 = round(total_score / skill_count,1)

        if Target.objects.filter(person=person,sub_cat=sub_category).count() > 0:
            sub_category.target = Target.objects.filter(person=person,sub_cat=sub_category)[0].target.level
            sub_category.save()

    return render(request, 'people_ind.html', {'list': list,'sub_categories':sub_categories,'person':person,'scores':scores,'colour':colour})


def people_skill_update(request, people_id, sub_category_id):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    list = Skill.objects.filter(sub_category=sub_category_id)
    sub_category = Skill_Cat.objects.filter(id=sub_category_id)[0]
    person = People.objects.get(id=people_id)
    levels = Skill_Level.objects.all()
    buttons = {1,2,3,4,5}
    for skill in list:
        if Score.objects.filter(person=person, skill=skill).count() > 0:
            result = Score.objects.filter(person=person, skill=skill)[0].score
            skill.score_temp = result
            skill.save()
        else:
            skill.score_temp = None

    if request.method == 'POST':
        for item in list:
            try:
                print(request.POST)
                score = request.POST.__getitem__(str(item.id))
                existing_query = Score.objects.filter(skill=item,person=person)
                if existing_query.count() == 0:
                    Score(skill=item,person=person,score=score).save()
                else:
                    existing = existing_query[0]
                    existing.score = score
                    existing.save()
            except:
                pass
        return redirect('/people_ind/' + people_id)
    return render(request, 'people_skill_update.html', {'list': list,'person':person,'sub_category': sub_category,
                                                        'levels':levels, 'buttons':buttons,'colour':colour})

def people_target_update(request,people_id,sub_category_id,level_id):
    sub_category = Skill_Cat.objects.filter(id=sub_category_id)[0]
    person = People.objects.get(id=people_id)
    levels = Skill_Level.objects.all()
    level = levels[int(level_id)]
    if Target.objects.filter(sub_cat=sub_category,person=person).count() == 0:
        new_target = Target(sub_cat=sub_category,person=person,target=level)
        new_target.save()
    else:
        target_obj = Target.objects.filter(sub_cat=sub_category,person=person)[0]
        target_obj.target = level
        target_obj.save()
    return redirect('/people_ind/' + people_id)

def people_update(request, id):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    item = People.objects.get(pk=id)
    form = PeopleForm(request.POST or None, instance=item)
    if form.is_valid():
        form.save()
        return redirect('people_list')
    return render(request, 'update.html', {'item': item, 'form': form,'colour':colour})

def people_delete(request, id):
    item = People.objects.get(pk=id)
    item.delete()
    return redirect('people_list')

def people_delete_all(request):
    item = People.objects.all()
    item.delete()
    return redirect('people_list')

def people_upload(request, id):
    file = File.objects.filter(id=id)[0]
    path = str(file.document.url)[1:]
    try:
        wb = xl.load_workbook(path)
    except:
        list = People.objects.all().order_by('name')
        error = True
        return render(request, 'people_list.html', {'list': list, 'error': error})
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        manager = sheet.cell(row, 2).value
        role = sheet.cell(row, 3).value

        if People.objects.filter(name=manager).count() > 0:
            parent_object = People.objects.filter(name=manager)[0]
        else:
            parent_object = None
        if name is not None:
            People(name=name, manager=parent_object, role=role).save()
    return redirect('people_list')

def skill_list(request):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    list = Skill.objects.all()
    rows = Skill_Cat.objects.all()
    levels = Skill_Level.objects.all()
    return render(request, 'skill_list.html', {'list': list,'rows':rows,'levels':levels,'colour':colour})

def level_list(request):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    list = Skill_Level.objects.all()
    list2 = Role_Level.objects.all()
    return render(request, 'level_list.html', {'list': list,'list2': list2,'colour':colour})

def skill_ind(request, sub_category_id, level_id):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    list = Skill.objects.filter(sub_category=sub_category_id)
    sub_category = Skill_Cat.objects.filter(id=sub_category_id)[0]
    buttons = {1,2,3,4,5}
    if request.method == 'POST':
        return redirect('skill_list')
    return render(request, 'skill_ind.html', {'list': list,'sub_category': sub_category,'buttons':buttons,'colour':colour})

def skill_update(request, id):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    item = Skill.objects.get(pk=id)
    form = SkillForm(request.POST or None, instance=item)
    if form.is_valid():
        form.save()
        return redirect('skill_list')
    return render(request, 'update.html', {'item': item, 'form': form,'colour':colour})

def skill_delete(request, id):
    item = Skill.objects.get(pk=id)
    item.delete()
    return redirect('skill_list')

def skill_delete_all(request):
    item = Skill.objects.all()
    item.delete()
    item = Skill_Level.objects.all()
    item.delete()
    item = Skill_Cat.objects.all()
    item.delete()
    return redirect('skill_list')

def skill_cat_upload(request, id):
    file = File.objects.filter(id=id)[0]
    path = str(file.document.url)[1:]
    try:
        wb = xl.load_workbook(path)
    except:
        list = Skill.objects.all().order_by('category')
        error = True
        return render(request, 'skill_list.html', {'list': list, 'error': error})
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        category = sheet.cell(row, 1).value
        sub_category = sheet.cell(row, 2).value
        if Skill_Cat.objects.all().filter(category=category).filter(sub_category=sub_category).count() == 0 and sub_category is not None:
            Skill_Cat(category=category, sub_category=sub_category).save()
    return redirect('skill_list')

def level_upload(request, id):
    file = File.objects.filter(id=id)[0]
    path = str(file.document.url)[1:]
    try:
        wb = xl.load_workbook(path)
    except:
        list = Skill_Level.objects.all()
        list2 = Role_Level.objects.all()
        error = True
        return render(request, 'skill_list.html', {'list': list, 'list2': list2, 'error': error})
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        level = sheet.cell(row, 1).value
        if level == None:
            level = "No level"
        if Skill_Level.objects.all().filter(level=level).count() == 0 and level is not None:
            level_obj = Skill_Level(level=level)
            level_obj.save()
    return redirect('level_list')

def role_upload(request, id):
    file = File.objects.filter(id=id)[0]
    path = str(file.document.url)[1:]
    try:
        wb = xl.load_workbook(path)
    except:
        list = Skill_Level.objects.all()
        list2 = Role_Level.objects.all()
        error = True
        return render(request, 'skill_list.html', {'list': list, 'list2': list2, 'error': error})
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        role = sheet.cell(row, 1).value
        if role == None:
            role = "No level"
        if Role_Level.objects.all().filter(role_level=role).count() == 0 and role is not None:
            role_obj = Role_Level(role_level=role)
            role_obj.save()
    return redirect('level_list')

def level_delete_all(request):
    item = Skill_Level.objects.all()
    item.delete()
    return redirect('level_list')

def role_delete_all(request):
    Role_Level.objects.all().delete()
    return redirect('level_list')

def skill_upload(request, id):
    file = File.objects.filter(id=id)[0]
    path = str(file.document.url)[1:]
    try:
        wb = xl.load_workbook(path)
    except:
        list = Skill.objects.all().order_by('category')
        error = True
        return render(request, 'skill_list.html', {'list': list, 'error': error})

    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        category = sheet.cell(row, 1).value
        sub_category = sheet.cell(row, 2).value
        role_level = sheet.cell(row, 3).value
        question = sheet.cell(row, 4).value
        if Skill_Cat.objects.all().filter(category=category, sub_category=sub_category).count() > 0:
            sub_category_obj = Skill_Cat.objects.all().filter(category=category, sub_category=sub_category)[0]
            if Role_Level.objects.all().filter(role_level=role_level).count() > 0:
                role_level_obj = Role_Level.objects.all().filter(role_level=role_level)[0]
                if Skill.objects.all().filter(sub_category=sub_category_obj,question=question).count() == 0:
                    Skill(sub_category=sub_category_obj, question=question, role_level=role_level_obj).save()
            else:
                print(f"{role_level} not found")
        else:
            print(f"{sub_category} not found")
    return redirect('skill_list')

def colour_list(request):
    list = Colour.objects.all()
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()
    return render(request, 'colour_list.html', {'list': list,'colour':colour})

def colour_activate(request, id):
    active = Colour.objects.filter(active=True)
    for item in active:
        item.active = False
        item.save()
    new_active = Colour.objects.filter(id=id)[0]
    new_active.active = True
    new_active.save()
    return redirect('colour_list')

def colour_delete_all(request):
    item = Colour.objects.all()
    item.delete()
    return redirect('colour_list')

def colour_upload(request, id):
    file = File.objects.filter(id=id)[0]
    path = str(file.document.url)[1:]
    try:
        wb = xl.load_workbook(path)
    except:
        list = Colour.objects.all().order_by('name')
        error = True
        return render(request, 'colour_list.html', {'list': list, 'error': error})
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        active = sheet.cell(row, 2).value
        nav = sheet.cell(row, 3).value
        primary_dark = sheet.cell(row, 4).value
        primary_light = sheet.cell(row, 5).value
        secondary_dark = sheet.cell(row, 6).value
        secondary_light = sheet.cell(row, 7).value

        if name is not None:
            Colour(name=name, active=active, nav=nav, primary_dark=primary_dark, primary_light=primary_light,
                   secondary_dark=secondary_dark, secondary_light=secondary_light).save()
    return redirect('colour_list')

def file_list(request):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    list = File.objects.all().order_by('name')
    return render(request, 'file_list.html', {'list': list,'colour':colour})

def file_upload(request):
    if Colour.objects.filter(active=True).count() > 0:
        colour = Colour.objects.filter(active=True)[0]
    elif Colour.objects.all().count() > 0:
        colour = Colour.objects.all()[0]
    else:
        colour = Colour().save()

    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('file_list')
    else:
        form = FileForm()
    return render(request, 'file_upload.html',{'form':form,'colour':colour})

def file_delete(request, file_id):
    if request.method == 'POST':
        file = File.objects.get(pk=file_id)
        file.delete()
    return redirect('file_list')

